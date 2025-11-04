#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

"""
@File    : ui_processing.py
@Author  : Link
@Time    : 2022/7/31 13:35
@Mark    : 
"""

from typing import Union

import numpy as np
from PySide2.QtGui import QStandardItemModel, QStandardItem, Qt
from PySide2.QtWidgets import QWidget
from PySide2.QtCore import Slot, QModelIndex

from ui_component.ui_app_variable import UiGlobalVariable
from ui_component.ui_analysis_stdf.ui_designer.ui_processing import Ui_Form

import pyqtgraph as pg
import pandas as pd


class ProcessWidget(QWidget, Ui_Form):
    """
    需要的时候再运行, 即功能打开时运行而不是数据一改变就运行
    """
    select_item_list = QStandardItemModel()
    top_item_list = QStandardItemModel()  # yield, avg, limit ...
    bot_item_list = QStandardItemModel()  # group by item
    li = None

    def __init__(self, parent=None, icon=None):
        super(ProcessWidget, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle("制程能力")
        if icon:
            self.setWindowIcon(icon)
        self.listView_3.setModel(self.select_item_list)
        self.listView_2.setModel(self.top_item_list)
        self.listView_2.clicked.connect(self.top_row_change)
        self.listView.setModel(self.bot_item_list)
        self.listView.clicked.connect(self.bot_row_change)
        self.cpk_info_table = pg.TableWidget(self)
        self.verticalLayout.addWidget(self.cpk_info_table)
        self.cpk_info_table.setEditable(True)
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 15)

        self.init_listView_3()
        self.init_listView_2()

        # 连接导出按钮
        self.pushButton.clicked.connect(self.on_export_excel_clicked)

        # 添加增强报告按钮连接
        if hasattr(self, 'pushButton_6'):
            self.pushButton_6.clicked.connect(self.on_export_enhanced_report_clicked)

        # 添加制程能力图表生成按钮
        if hasattr(self, 'pushButton_7'):
            self.pushButton_7.clicked.connect(self.on_generate_capability_charts_clicked)

    def init_listView_3(self):
        self.select_item_list.clear()
        for index, each in enumerate(UiGlobalVariable.PROCESS_VALUE):
            item = QStandardItem(each)
            item.setCheckState(Qt.Unchecked)
            item.setCheckable(True)
            if index == 0:
                item.setCheckState(Qt.Checked)
            self.select_item_list.appendRow(item)

    def get_listView_3_choose_items(self) -> Union[list, None]:
        select_item_list = []
        for index, each in enumerate(UiGlobalVariable.PROCESS_VALUE):
            temp = self.select_item_list.item(index)
            if temp.checkState() == Qt.Checked:
                select_item_list.append(temp.text())
        return select_item_list if select_item_list else None

    def init_listView_2(self):
        self.top_item_list.clear()
        for index, each in enumerate(UiGlobalVariable.PROCESS_TOP_ITEM_LIST):
            item = QStandardItem(each)
            self.top_item_list.appendRow(item)

    def gen_listView(self):
        """
        通过 GROUP 和 DA_GROUP 来做处理
        :return:
        """
        bot_item_list_df = self.li.front_df["GROUP"] + "_" + self.li.front_df["DA_GROUP"]  # type:pd.DataFrame
        bot_item_list = bot_item_list_df.drop_duplicates(keep="first").tolist()
        self.bot_item_list.clear()
        for index, each in enumerate(bot_item_list):
            item = QStandardItem(each)
            self.bot_item_list.appendRow(item)

    def set_data(self, li):
        if self.li is not None:
            self.li.front_df_signal.disconnect(self.set_front_df_process)
        self.li = li
        self.li.front_df_signal.connect(self.set_front_df_process)

    @Slot()
    def set_front_df_process(self):
        """
        将数据显示到前台中. 一般在这里的时候, 数据已经经理了group by阶段了
        :return:
        """
        if self.li is None:
            return False
        self.gen_listView()

    @Slot(QModelIndex)
    def top_row_change(self, model_index: QModelIndex):
        """
        放良率. 和 avg 对比
        :param model_index:
        :return:
        """
        if self.li is None:
            return

        if "GROUP" not in self.li.front_df or "DA_GROUP" not in self.li.front_df:
            return

        if model_index.data() == "yield":
            df_group = self.li.front_df.groupby(["GROUP", "DA_GROUP"])
            yield_data = []  # type:list
            for key, each_df in df_group:
                if not isinstance(key, tuple):
                    item_name = str(key)
                else:
                    item_name = '-'.join([str(ea) for ea in key])
                total = len(each_df)
                fail_num = len(each_df[each_df["FAIL_FLAG"] != 1])
                pass_num = len(each_df) - fail_num
                yield_data.append({
                    "Item": item_name,
                    "Total": total,
                    "Pass": pass_num,
                    "Fail": fail_num,
                    "Yield": "{}%".format(round(pass_num / total * 100, 3)),
                })
            self.cpk_info_table.setData(yield_data)
            return

        if model_index.data() == "data":
            """
            这个稍微复杂一些, 先将数据都获取到, 然后再整理起来
            """
            item_list = self.get_listView_3_choose_items()
            if item_list is None:
                return
            df_group = self.li.front_df.groupby(["GROUP", "DA_GROUP"])
            group_cpk_dict = dict()
            for key, each_df in df_group:
                if not isinstance(key, tuple):
                    item_name = str(key)
                else:
                    item_name = '-'.join([str(ea) for ea in key])
                _cpk_df = each_df[each_df["FAIL_FLAG"] == 1][self.li.front_limit_dict.keys()]
                _mean = _cpk_df.mean()
                _std = _cpk_df.std()
                temp_data_list = []
                for index, item_key in enumerate(self.li.front_limit_dict.keys()):
                    item = self.li.front_limit_dict[item_key]
                    temp_std = _std[item_key]
                    temp_mean = _mean[item_key]
                    if temp_std == 0:
                        cpk = 0
                    else:
                        cpk = round(min(
                            [(item.h_limit - temp_mean) / (3 * temp_std),
                             (temp_mean - item.l_limit) / (3 * temp_std)])
                            , 6)
                    temp_dict = {
                        "SORT": item.test_sort,
                        "TEST_TYPE": item.test_type,
                        "Text": item_key,
                        "UNITS": item.unit,
                        "LO_LIMIT": item.l_limit,
                        "HI_LIMIT": item.h_limit,
                        "LO_LIMIT_TYPE": item.l_limit_type,
                        "HI_LIMIT_TYPE": item.h_limit_type,
                        f"{item_name}_Avg": round(_mean[item_key], 5),
                        f"{item_name}_Stdev": round(_std[item_key], 5),
                        f"{item_name}_Cpk": cpk,
                    }
                    temp_data_list.append(temp_dict)
                group_cpk_dict[item_name] = temp_data_list

            """
            前台展示数据
            """
            data_table_list = []
            for index, key in enumerate(group_cpk_dict.keys()):
                temp_each_data = group_cpk_dict[key]
                if index == 0:
                    for i, row in enumerate(temp_each_data):
                        d = {
                            "SORT": row["SORT"],
                            "TEST_TYPE": row["TEST_TYPE"],
                            "Text": row["Text"],
                            "UNITS": row["UNITS"],
                            "LO_LIMIT": row["LO_LIMIT"],
                            "HI_LIMIT": row["HI_LIMIT"],
                            "LO_LIMIT_TYPE": row["LO_LIMIT_TYPE"],
                            "HI_LIMIT_TYPE": row["HI_LIMIT_TYPE"],
                        }
                        for each in item_list:
                            if each == "mean":
                                d[f"{key}_Avg"] = row[f"{key}_Avg"]
                            if each == "std":
                                d[f"{key}_Stdev"] = row[f"{key}_Stdev"]
                            if each == "cpk":
                                d[f"{key}_Cpk"] = row[f"{key}_Cpk"]
                        data_table_list.append(d)
                else:
                    for i, row in enumerate(temp_each_data):
                        for each in item_list:
                            if each == "mean":
                                data_table_list[i][f"{key}_Avg"] = row[f"{key}_Avg"]
                            if each == "std":
                                data_table_list[i][f"{key}_Stdev"] = row[f"{key}_Stdev"]
                            if each == "cpk":
                                data_table_list[i][f"{key}_Cpk"] = row[f"{key}_Cpk"]
            self.cpk_info_table.setData(data_table_list)
            return

    @Slot(QModelIndex)
    def bot_row_change(self, model_index: QModelIndex):
        """
        放单个数据, 单个数据可以有value类型和diff类型
        :param model_index:
        :return:
        """
        if self.radioButton_2.isChecked():
            group, da_group = model_index.data().split("_")
            df = self.li.front_df[(self.li.front_df["GROUP"] == group) & (self.li.front_df["DA_GROUP"] == da_group)]

            cpk_list = []
            first_fail_dict = {}
            for num in df[df["FAIL_FLAG"] != 1]["FIRST_FAIL"]:
                if num not in first_fail_dict:
                    first_fail_dict[num] = 1
                    continue
                first_fail_dict[num] += 1

            temp_df = df[df["FAIL_FLAG"] == 1][self.li.front_limit_dict.keys()]
            _mean, _min, _max, _std, _median = temp_df.mean(), temp_df.min(), temp_df.max(), temp_df.std(), temp_df.median()
            for key, item in self.li.front_limit_dict.items():
                try:
                    test_num, test_txt = key.split(":", 1)
                except ValueError:
                    raise Exception("重大错误: 测试数据测试项目没有指定TEST_NO和TEST_TEXT,测试程序漏洞@!!!!!")
                reject_qty = 0
                logic_or = []
                if item.l_limit_type == ">":
                    logic_or.append((df[key] <= item.l_limit))
                if item.l_limit_type == ">=":
                    logic_or.append((df[key] < item.l_limit))
                if item.h_limit_type == "<":
                    logic_or.append((df[key] >= item.h_limit))
                if item.h_limit_type == "<=":
                    logic_or.append((df[key] > item.h_limit))
                if item.h_limit_type == "=":
                    logic_or.append((df[key] != item.h_limit))
                if len(logic_or) == 1:
                    items = logic_or[0]
                    reject_qty = len(df.loc[items])
                if len(logic_or) > 1:
                    items = np.logical_or(*logic_or)
                    reject_qty = len(df.loc[items])

                fail_qty = first_fail_dict.get(item.test_num, 0)
                temp_std, temp_mean = _std[key], _mean[key]
                cpk = 0 if temp_std == 0 else round(min([(item.h_limit - temp_mean) / (3 * temp_std),
                                                         (temp_mean - item.l_limit) / (3 * temp_std)]), 6)
                temp_dict = {
                    "SORT": item.test_sort,
                    "TEST_TYPE": item.test_type,
                    "TEST_NUM": test_num,
                    "TEST_TEXT": test_txt,
                    "UNITS": str(item.unit),
                    "LO_LIMIT": item.l_limit,
                    "HI_LIMIT": item.h_limit,
                    "Average": round(_mean[key], 6),
                    "Stdev": round(_std[key], 6),
                    "Cpk": cpk,
                    "Text": key,
                    "Total": len(df),
                    "Fail": fail_qty,
                    "Fail/Total": "{}%".format(round(fail_qty / len(df) * 100, 3)),
                    "Reject": reject_qty,
                    "Reject/Total": "{}%".format(round(reject_qty / len(df) * 100, 3)),
                    "Min": round(_min[key], 6),
                    "Max": round(_max[key], 6),
                    "LO_LIMIT_TYPE": item.l_limit_type,
                    "HI_LIMIT_TYPE": item.h_limit_type,
                }
                cpk_list.append(temp_dict)
            self.cpk_info_table.setData(cpk_list)
            return
        if self.radioButton.isChecked():
            """
            DIFF, 只看均值以及PTR项目
            """
            group, da_group = model_index.data().split("_")
            diff_compare_df = self.li.front_df[
                (self.li.front_df["GROUP"] == group) & (self.li.front_df["DA_GROUP"] == da_group)]
            length = len(diff_compare_df)
            start, stop = int(length * 0.05), int(length * 0.95)

            diff_data_list = []
            df_group = self.li.front_df.groupby(["GROUP", "DA_GROUP"])
            for test_item, test_tuple in self.li.front_limit_dict.items():  # type:str, GlobalVariable.LimitClass
                if test_tuple.test_type == "FTR":
                    continue
                temp_dict = dict()
                diff_data_list.append(temp_dict)
                temp_dict["SORT"] = test_tuple.test_sort
                temp_dict["ITEM"] = test_item
                temp_dict["Unit"] = test_tuple.unit
                temp_dict["LO_LIMIT"] = test_tuple.l_limit
                temp_dict["HI_LIMIT"] = test_tuple.h_limit
                temp_dict["LO_LIMIT_TYPE"] = test_tuple.l_limit_type
                temp_dict["HI_LIMIT_TYPE"] = test_tuple.h_limit_type

                compare_data = np.mean(sorted(diff_compare_df[test_item].to_list())[start: stop])
                for key, each_df in df_group:
                    if not isinstance(key, tuple):
                        item_name = str(key)
                    else:
                        item_name = '-'.join([str(ea) for ea in key])
                    temp_data = each_df[test_item].to_list()
                    temp_length = len(each_df)
                    temp_start, temp_stop = int(temp_length * 0.05), int(temp_length * 0.95)
                    temp_data = sorted(temp_data)[temp_start: temp_stop]
                    temp_data = np.mean(temp_data)
                    gap = compare_data - temp_data
                    temp_dict[item_name] = gap

            self.cpk_info_table.setData(diff_data_list)
            return

    @Slot()
    def on_export_excel_clicked(self):
        """导出制程能力报告为Excel"""
        if self.li is None or not hasattr(self.li, 'capability_key_list') or not self.li.capability_key_list:
            from PySide2.QtWidgets import QMessageBox
            QMessageBox.warning(self, "警告", "没有制程能力数据可以导出！\n请先计算制程能力数据。")
            return

        from PySide2.QtWidgets import QFileDialog, QMessageBox
        from common.app_variable import GlobalVariable
        import os
        import pandas as pd

        # 确保C:\1_STDF目录存在
        stdf_dir = "C:/1_STDF"
        os.makedirs(stdf_dir, exist_ok=True)

        # 生成带时间戳的文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"制程能力报告_{timestamp}.xlsx"

        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出制程能力报告",
            os.path.join(stdf_dir, default_filename),
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        try:
            self.export_capability_report(file_path)
            QMessageBox.information(self, "成功", f"制程能力报告已导出到:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出制程能力报告失败:\n{str(e)}")

    def export_capability_report(self, file_path):
        """导出制程能力报告到Excel文件"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        # 创建Excel工作簿
        wb = Workbook()

        # 删除默认工作表
        wb.remove(wb.active)

        # 1. 制程能力汇总表
        if self.li.capability_key_list:
            capability_df = pd.DataFrame(self.li.capability_key_list)
            ws_capability = wb.create_sheet("制程能力汇总")

            # 添加标题
            ws_capability['A1'] = "制程能力分析报告"
            ws_capability['A1'].font = Font(size=16, bold=True)
            ws_capability.merge_cells('A1:H1')

            # 添加数据
            for r in dataframe_to_rows(capability_df, index=False, header=True):
                ws_capability.append(r)

            # 设置样式
            self._format_excel_sheet(ws_capability, capability_df)

        # 2. Top Fail分析表
        if self.li.top_fail_dict:
            top_fail_df = pd.DataFrame([
                {
                    'TEST_ID': k,
                    'FAIL_COUNT': v,
                    'TEST_NAME': self.li.get_text_by_test_id(k) if hasattr(self.li, 'get_text_by_test_id') else f"Test_{k}"
                }
                for k, v in self.li.top_fail_dict.items()
            ])
            # 按失效数量排序
            top_fail_df = top_fail_df.sort_values('FAIL_COUNT', ascending=False)

            ws_top_fail = wb.create_sheet("Top_Fail分析")
            ws_top_fail['A1'] = "Top Fail分析"
            ws_top_fail['A1'].font = Font(size=14, bold=True)

            for r in dataframe_to_rows(top_fail_df, index=False, header=True):
                ws_top_fail.append(r)

            self._format_excel_sheet(ws_top_fail, top_fail_df)

        # 3. 数据汇总表
        if self.li.select_summary is not None:
            ws_summary = wb.create_sheet("数据汇总")
            ws_summary['A1'] = "数据汇总信息"
            ws_summary['A1'].font = Font(size=14, bold=True)

            for r in dataframe_to_rows(self.li.select_summary, index=False, header=True):
                ws_summary.append(r)

            self._format_excel_sheet(ws_summary, self.li.select_summary)

        # 4. 如果有当前表格数据，也导出
        current_data = self.cpk_info_table.data
        if current_data:
            current_df = pd.DataFrame(current_data)
            ws_current = wb.create_sheet("当前分析结果")
            ws_current['A1'] = "当前分析结果"
            ws_current['A1'].font = Font(size=14, bold=True)

            for r in dataframe_to_rows(current_df, index=False, header=True):
                ws_current.append(r)

            self._format_excel_sheet(ws_current, current_df)

        # 保存文件
        wb.save(file_path)

    def _format_excel_sheet(self, worksheet, dataframe):
        """格式化Excel工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 设置标题样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 格式化标题行
        for cell in worksheet[3]:  # 假设数据从第3行开始
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # 格式化数据行
        for row in worksheet.iter_rows(min_row=4, max_row=len(dataframe)+3):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # 自动调整列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    @Slot()
    def on_export_enhanced_report_clicked(self):
        """导出增强制程能力报告（Excel + PDF + 分布图）"""
        if not self.li or not self.li.capability_key_list:
            QMessageBox.warning(self, "警告", "没有制程能力数据可导出！")
            return

        try:
            from temp.enhanced_capability_report import EnhancedCapabilityReporter

            # 选择输出目录
            output_dir = QFileDialog.getExistingDirectory(
                self,
                "选择报告输出目录",
                "C:/1_STDF/CAPABILITY_REPORTS"
            )

            if not output_dir:
                return

            # 创建进度对话框
            progress = QProgressDialog("正在生成增强制程能力报告...", "取消", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)
            progress.show()

            # 更新进度
            progress.setValue(10)
            QApplication.processEvents()

            # 创建报告生成器
            reporter = EnhancedCapabilityReporter()

            progress.setValue(30)
            QApplication.processEvents()

            # 检查是否有DTP数据
            if not hasattr(self.li, 'df_module') or self.li.df_module is None:
                QMessageBox.warning(self, "警告", "没有找到测试数据！")
                return

            dtp_df = self.li.df_module.dtp_df

            progress.setValue(50)
            QApplication.processEvents()

            # 生成报告
            results = reporter.generate_complete_report(
                self.li.capability_key_list,
                dtp_df,
                output_dir
            )

            progress.setValue(90)
            QApplication.processEvents()

            # 清理临时文件
            reporter.cleanup_temp_files()

            progress.setValue(100)
            QApplication.processEvents()

            # 显示结果
            if results:
                message = "增强制程能力报告生成成功！\n\n生成的文件：\n"
                for report_type, file_path in results.items():
                    message += f"• {report_type.upper()}: {os.path.basename(file_path)}\n"

                message += f"\n输出目录: {output_dir}"

                reply = QMessageBox.information(
                    self,
                    "报告生成成功",
                    message,
                    QMessageBox.Ok | QMessageBox.Open
                )

                # 如果用户选择打开，则打开输出目录
                if reply == QMessageBox.Open:
                    import subprocess
                    subprocess.Popen(f'explorer "{output_dir}"')
            else:
                QMessageBox.warning(self, "错误", "报告生成失败！请检查数据和权限。")

        except ImportError:
            QMessageBox.critical(
                self,
                "模块错误",
                "无法导入增强报告模块！\n请确保 temp/enhanced_capability_report.py 文件存在。"
            )
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成增强报告时发生错误：\n{str(e)}")

    def export_enhanced_capability_report_to_path(self, output_dir: str) -> bool:
        """
        导出增强制程能力报告到指定路径
        供其他模块调用的接口
        """
        try:
            from temp.enhanced_capability_report import EnhancedCapabilityReporter

            if not self.li or not self.li.capability_key_list:
                return False

            if not hasattr(self.li, 'df_module') or self.li.df_module is None:
                return False

            reporter = EnhancedCapabilityReporter()
            results = reporter.generate_complete_report(
                self.li.capability_key_list,
                self.li.df_module.dtp_df,
                output_dir
            )

            reporter.cleanup_temp_files()
            return len(results) > 0

        except Exception as e:
            print(f"导出增强报告失败: {e}")
            return False

    @Slot()
    def on_generate_capability_charts_clicked(self):
        """生成制程能力图表"""
        try:
            from PySide2.QtWidgets import QMessageBox, QProgressDialog, QApplication
            from PySide2.QtCore import Qt

            # 检查matplotlib
            try:
                import matplotlib.pyplot as plt
                import matplotlib
                matplotlib.use('Agg')  # 使用非交互式后端
                import numpy as np
                print("matplotlib导入成功")
            except ImportError:
                QMessageBox.critical(self, "模块错误", "matplotlib模块未安装！\n请安装matplotlib: pip install matplotlib")
                return

            import os

            # 检查数据
            if not self.li:
                QMessageBox.warning(self, "警告", "没有数据接口！")
                return

            if not hasattr(self.li, 'capability_key_list') or not self.li.capability_key_list:
                QMessageBox.warning(self, "警告", "没有制程能力数据可以生成图表！\n请先计算制程能力数据。")
                return

            print(f"找到制程能力数据: {len(self.li.capability_key_list)}项")

            # 确保输出目录存在
            output_dir = "C:/1_STDF/CAPABILITY_CHARTS"
            os.makedirs(output_dir, exist_ok=True)

            # 创建进度对话框
            progress = QProgressDialog("正在生成制程能力图表...", "取消", 0, 100, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.setAutoClose(True)
            progress.setAutoReset(True)
            progress.show()

            # 生成时间戳
            import datetime
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

            # 获取制程能力数据
            capability_data = self.li.capability_key_list
            total_items = len(capability_data)

            if total_items == 0:
                QMessageBox.warning(self, "警告", "没有有效的制程能力数据！")
                return

            progress.setValue(10)
            QApplication.processEvents()

            # 生成CPK分布图
            self._generate_cpk_distribution_chart(capability_data, output_dir, timestamp)
            progress.setValue(30)
            QApplication.processEvents()

            # 生成CPK趋势图
            self._generate_cpk_trend_chart(capability_data, output_dir, timestamp)
            progress.setValue(50)
            QApplication.processEvents()

            # 生成制程能力汇总图
            self._generate_capability_summary_chart(capability_data, output_dir, timestamp)
            progress.setValue(70)
            QApplication.processEvents()

            # 生成Top 10 低CPK项目图
            self._generate_top_low_cpk_chart(capability_data, output_dir, timestamp)
            progress.setValue(90)
            QApplication.processEvents()

            progress.setValue(100)
            QApplication.processEvents()

            # 显示结果
            message = f"制程能力图表生成成功！\n\n生成的图表：\n"
            message += f"• CPK分布图: CPK分布图_{timestamp}.png\n"
            message += f"• CPK趋势图: CPK趋势图_{timestamp}.png\n"
            message += f"• 制程能力汇总图: 制程能力汇总_{timestamp}.png\n"
            message += f"• Top10低CPK项目图: Top10低CPK_{timestamp}.png\n"
            message += f"\n输出目录: {output_dir}"

            reply = QMessageBox.information(
                self,
                "图表生成成功",
                message,
                QMessageBox.Ok | QMessageBox.Open
            )

            if reply == QMessageBox.Open:
                import subprocess
                subprocess.Popen(f'explorer "{output_dir}"')

        except ImportError:
            QMessageBox.critical(self, "模块错误", "matplotlib模块未安装！\n请安装matplotlib: pip install matplotlib")
        except Exception as e:
            QMessageBox.critical(self, "生成失败", f"制程能力图表生成失败:\n{str(e)}")

    def _generate_cpk_distribution_chart(self, capability_data, output_dir, timestamp):
        """生成CPK分布图"""
        try:
            import matplotlib.pyplot as plt
            import numpy as np

            # 提取CPK值
            cpk_values = []
            for item in capability_data:
                cpk = item.get('CPK', 0)
                if cpk is not None and not np.isnan(cpk) and cpk != 0:
                    cpk_values.append(cpk)

            if not cpk_values:
                print("没有有效的CPK值用于生成分布图")
                return

            print(f"生成CPK分布图，有效CPK值: {len(cpk_values)}个")

        plt.figure(figsize=(10, 6))
        plt.hist(cpk_values, bins=20, alpha=0.7, color='skyblue', edgecolor='black')
        plt.axvline(x=1.33, color='red', linestyle='--', label='CPK=1.33 (可接受)')
        plt.axvline(x=1.67, color='orange', linestyle='--', label='CPK=1.67 (良好)')
        plt.axvline(x=2.0, color='green', linestyle='--', label='CPK=2.0 (优秀)')

        plt.xlabel('CPK值')
        plt.ylabel('频次')
        plt.title('制程能力(CPK)分布图')
        plt.legend()
        plt.grid(True, alpha=0.3)

        # 添加统计信息
        mean_cpk = np.mean(cpk_values)
        median_cpk = np.median(cpk_values)
        plt.text(0.02, 0.98, f'平均CPK: {mean_cpk:.3f}\n中位数CPK: {median_cpk:.3f}\n样本数: {len(cpk_values)}',
                transform=plt.gca().transAxes, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8))

        plt.tight_layout()
        plt.savefig(f"{output_dir}/CPK分布图_{timestamp}.png", dpi=300, bbox_inches='tight')
        plt.close()

    def _generate_cpk_trend_chart(self, capability_data, output_dir, timestamp):
        """生成CPK趋势图"""
        import matplotlib.pyplot as plt
        import numpy as np

        # 按TEST_ID排序
        sorted_data = sorted(capability_data, key=lambda x: x.get('TEST_ID', 0))

        test_ids = []
        cpk_values = []
        test_names = []

        for item in sorted_data:
            test_id = item.get('TEST_ID', 0)
            cpk = item.get('CPK', 0)
            test_name = item.get('TEST_TXT', f'Test_{test_id}')

            if not np.isnan(cpk):
                test_ids.append(test_id)
                cpk_values.append(cpk)
                test_names.append(test_name)

        if not cpk_values:
            return

        plt.figure(figsize=(15, 8))
        colors = ['red' if cpk < 1.33 else 'orange' if cpk < 1.67 else 'green' for cpk in cpk_values]

        plt.scatter(range(len(cpk_values)), cpk_values, c=colors, alpha=0.7, s=50)
        plt.plot(range(len(cpk_values)), cpk_values, 'b-', alpha=0.5, linewidth=1)

        plt.axhline(y=1.33, color='red', linestyle='--', alpha=0.7, label='CPK=1.33 (可接受)')
        plt.axhline(y=1.67, color='orange', linestyle='--', alpha=0.7, label='CPK=1.67 (良好)')
        plt.axhline(y=2.0, color='green', linestyle='--', alpha=0.7, label='CPK=2.0 (优秀)')

        plt.xlabel('测试项目序号')
        plt.ylabel('CPK值')
        plt.title('制程能力(CPK)趋势图')
        plt.legend()
        plt.grid(True, alpha=0.3)

        # 设置x轴标签（每10个显示一个）
        step = max(1, len(test_ids) // 20)
        plt.xticks(range(0, len(test_ids), step),
                  [f'T{test_ids[i]}' for i in range(0, len(test_ids), step)],
                  rotation=45)

        plt.tight_layout()
        plt.savefig(f"{output_dir}/CPK趋势图_{timestamp}.png", dpi=300, bbox_inches='tight')
        plt.close()

    def _generate_capability_summary_chart(self, capability_data, output_dir, timestamp):
        """生成制程能力汇总图"""
        import matplotlib.pyplot as plt
        import numpy as np

        # 统计不同CPK等级的数量
        excellent = sum(1 for item in capability_data if item.get('CPK', 0) >= 2.0 and not np.isnan(item.get('CPK', 0)))
        good = sum(1 for item in capability_data if 1.67 <= item.get('CPK', 0) < 2.0 and not np.isnan(item.get('CPK', 0)))
        acceptable = sum(1 for item in capability_data if 1.33 <= item.get('CPK', 0) < 1.67 and not np.isnan(item.get('CPK', 0)))
        need_improve = sum(1 for item in capability_data if 1.0 <= item.get('CPK', 0) < 1.33 and not np.isnan(item.get('CPK', 0)))
        unqualified = sum(1 for item in capability_data if 0 < item.get('CPK', 0) < 1.0 and not np.isnan(item.get('CPK', 0)))
        invalid = sum(1 for item in capability_data if np.isnan(item.get('CPK', 0)) or item.get('CPK', 0) <= 0)

        categories = ['优秀\n(≥2.0)', '良好\n(1.67-2.0)', '可接受\n(1.33-1.67)',
                     '需改进\n(1.0-1.33)', '不合格\n(<1.0)', '无效\n(NaN/≤0)']
        counts = [excellent, good, acceptable, need_improve, unqualified, invalid]
        colors = ['green', 'lightgreen', 'yellow', 'orange', 'red', 'gray']

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))

        # 柱状图
        bars = ax1.bar(categories, counts, color=colors, alpha=0.7, edgecolor='black')
        ax1.set_ylabel('测试项目数量')
        ax1.set_title('制程能力等级分布')
        ax1.grid(True, alpha=0.3)

        # 在柱状图上添加数值标签
        for bar, count in zip(bars, counts):
            if count > 0:
                ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                        str(count), ha='center', va='bottom', fontweight='bold')

        # 饼图
        valid_counts = [c for c in counts if c > 0]
        valid_categories = [cat for cat, c in zip(categories, counts) if c > 0]
        valid_colors = [color for color, c in zip(colors, counts) if c > 0]

        if valid_counts:
            ax2.pie(valid_counts, labels=valid_categories, colors=valid_colors, autopct='%1.1f%%', startangle=90)
            ax2.set_title('制程能力等级占比')

        plt.tight_layout()
        plt.savefig(f"{output_dir}/制程能力汇总_{timestamp}.png", dpi=300, bbox_inches='tight')
        plt.close()

    def _generate_top_low_cpk_chart(self, capability_data, output_dir, timestamp):
        """生成Top 10 低CPK项目图"""
        import matplotlib.pyplot as plt
        import numpy as np

        # 过滤有效CPK值并排序
        valid_data = []
        for item in capability_data:
            cpk = item.get('CPK', 0)
            if not np.isnan(cpk) and cpk > 0:
                valid_data.append({
                    'TEST_ID': item.get('TEST_ID', 0),
                    'TEST_TXT': item.get('TEST_TXT', 'Unknown'),
                    'CPK': cpk,
                    'AVG': item.get('AVG', 0),
                    'STD': item.get('STD', 0)
                })

        # 按CPK排序，取最低的10个
        sorted_data = sorted(valid_data, key=lambda x: x['CPK'])
        top_10_low = sorted_data[:10]

        if not top_10_low:
            return

        plt.figure(figsize=(12, 8))

        test_names = [f"T{item['TEST_ID']}\n{item['TEST_TXT'][:15]}..." if len(item['TEST_TXT']) > 15
                     else f"T{item['TEST_ID']}\n{item['TEST_TXT']}" for item in top_10_low]
        cpk_values = [item['CPK'] for item in top_10_low]

        # 根据CPK值设置颜色
        colors = ['red' if cpk < 1.0 else 'orange' if cpk < 1.33 else 'yellow' for cpk in cpk_values]

        bars = plt.barh(range(len(test_names)), cpk_values, color=colors, alpha=0.7, edgecolor='black')

        # 添加参考线
        plt.axvline(x=1.0, color='red', linestyle='--', alpha=0.7, label='CPK=1.0')
        plt.axvline(x=1.33, color='orange', linestyle='--', alpha=0.7, label='CPK=1.33')
        plt.axvline(x=1.67, color='green', linestyle='--', alpha=0.7, label='CPK=1.67')

        plt.yticks(range(len(test_names)), test_names)
        plt.xlabel('CPK值')
        plt.title('Top 10 最低CPK测试项目')
        plt.legend()
        plt.grid(True, alpha=0.3, axis='x')

        # 在柱状图上添加CPK数值
        for i, (bar, cpk) in enumerate(zip(bars, cpk_values)):
            plt.text(bar.get_width() + 0.01, bar.get_y() + bar.get_height()/2,
                    f'{cpk:.3f}', ha='left', va='center', fontweight='bold')

        plt.tight_layout()
        plt.savefig(f"{output_dir}/Top10低CPK_{timestamp}.png", dpi=300, bbox_inches='tight')
        plt.close()
