#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

"""
@File    : utils.py
@Author  : Link
@Time    : 2022/7/31 20:15
@Mark    : 使用多进程来操作
"""
import os

import win32api
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

import pandas as pd

from common.app_variable import GlobalVariable
from common.func import tid_maker


class OpenXl:
    CenterAlign = Alignment(horizontal='center', vertical='center')
    RightAlign = Alignment(horizontal='right', vertical='center')
    LeftAlign = Alignment(horizontal='left', vertical='center', wrap_text=True)

    BlueFill = PatternFill('solid', fgColor='00B050')
    YellowFill = PatternFill('solid', fgColor='FFFF00')
    RedFill = PatternFill('solid', fgColor='FF0000')
    GreenFill = PatternFill('solid', fgColor='EBF1DE')

    CommentFont = Font(name="等线", size=9, color="FF0000")
    SpecialFont = Font(name="微软雅黑", size=9, color="0070C0")
    TitleFont = Font(name="微软雅黑", size=10, bold=True)
    ColumnFont = Font(name="等线", size=11, bold=True)
    TextFont = Font(name="等线", size=9)

    Thin = Side(border_style="thin", color="000000")
    TextBorder = Border(top=Thin, left=Thin, right=Thin, bottom=Thin)

    @staticmethod
    def excel_limit_run(summary_df: pd.DataFrame, limit_df: pd.DataFrame):
        """
        前四列固定, 从第二行开始写入
        """
        wb = Workbook()

        sheet_1 = wb.create_sheet("Limit")

        row_head = ["ID", "LOT_ID", "SBLOT_ID", "WAFER_ID", "TEST_COD", "FLOW_ID", "PART_TYP", "JOB_NAM"]
        # 检查列是否存在，如果不存在则使用可用的列
        available_columns = limit_df.columns.tolist()
        column_head = []

        # 必需的列
        if "TEST_ID" in available_columns:
            column_head.append("TEST_ID")

        # TEXT列可能不存在，使用TEST_TXT作为替代
        if "TEXT" in available_columns:
            column_head.append("TEXT")
        elif "TEST_TXT" in available_columns:
            column_head.append("TEST_TXT")
            # 重命名列以保持兼容性
            limit_df = limit_df.rename(columns={"TEST_TXT": "TEXT"})

        if "UNITS" in available_columns:
            column_head.append("UNITS")

        # 如果没有足够的列，使用所有可用列
        if len(column_head) < 2:
            column_head = ["TEST_ID", "TEST_TXT", "UNITS"] if "TEST_TXT" in available_columns else available_columns[:3]

        # 确保列名在DataFrame中存在
        existing_columns = [col for col in column_head if col in limit_df.columns]
        if not existing_columns:
            existing_columns = available_columns[:3]  # 使用前3列作为备选

        df_limit = limit_df[existing_columns].drop_duplicates(subset=None, keep='first')

        for row, each in enumerate(row_head):
            t_row = row + 1
            sheet_1.cell(t_row, 2).value = each
            sheet_1.cell(t_row, 2).font = OpenXl.TitleFont
            sheet_1.cell(t_row, 2).fill = OpenXl.BlueFill

        title_row = len(row_head) + 1
        head_row = title_row + 1

        now_column = 1
        # 使用实际存在的列名
        actual_columns = [col for col in existing_columns if col in df_limit.columns]

        for each in actual_columns:
            for i in range(len(df_limit)):
                limit_data = df_limit.iloc[i]
                # 安全地获取值，如果列不存在则使用空字符串
                try:
                    value = limit_data[each] if each in limit_data.index else ""
                except (KeyError, IndexError):
                    value = ""
                sheet_1.cell(head_row + i, now_column).value = value
                sheet_1.cell(head_row + i, now_column).font = OpenXl.SpecialFont
                sheet_1.cell(head_row + i, now_column).border = OpenXl.TextBorder

            # 使用友好的列名显示
            display_name = each
            if each == "TEST_TXT":
                display_name = "TEXT"

            sheet_1.cell(title_row, now_column).value = display_name
            sheet_1.cell(title_row, now_column).alignment = OpenXl.CenterAlign
            sheet_1.cell(title_row, now_column).font = OpenXl.SpecialFont
            sheet_1.cell(title_row, now_column).fill = OpenXl.YellowFill
            sheet_1.cell(title_row, now_column).border = OpenXl.TextBorder
            now_column += 1

        summary_info = summary_df.to_dict(orient='records')

        for index, lot in enumerate(summary_info):

            for row, each in enumerate(row_head):
                start_row = row + 1
                sheet_1.merge_cells(start_row=start_row, start_column=now_column, end_row=start_row,
                                    end_column=now_column + 1)
                sheet_1.cell(start_row, now_column).value = lot[each]
                sheet_1.cell(start_row, now_column).alignment = OpenXl.CenterAlign
                sheet_1.cell(start_row, now_column).font = OpenXl.TitleFont
                sheet_1.cell(start_row, now_column).fill = OpenXl.BlueFill
                # sheet_1.cell(start_row, now_column).border = OpenXl.TextBorder

            for each in ["LO_LIMIT", "HI_LIMIT"]:
                # 检查是否有ID列，如果没有则使用所有数据
                if "ID" in limit_df.columns:
                    temp_lot = limit_df[limit_df.ID == lot["ID"]]
                else:
                    temp_lot = limit_df.copy()

                # 检查是否有TEXT列用于索引
                if "TEXT" in temp_lot.columns:
                    temp_lot = temp_lot.set_index("TEXT")
                elif "TEST_TXT" in temp_lot.columns:
                    temp_lot = temp_lot.set_index("TEST_TXT")

                sheet_1.cell(title_row, now_column).value = each
                sheet_1.cell(title_row, now_column).alignment = OpenXl.CenterAlign
                sheet_1.cell(title_row, now_column).font = OpenXl.SpecialFont
                sheet_1.cell(title_row, now_column).fill = OpenXl.YellowFill
                sheet_1.cell(title_row, now_column).border = OpenXl.TextBorder

                for i in range(len(df_limit)):
                    limit_data = df_limit.iloc[i]

                    # 获取TEXT字段，可能是TEXT或TEST_TXT
                    text_value = None
                    if "TEXT" in limit_data.index:
                        text_value = limit_data.TEXT
                    elif "TEST_TXT" in limit_data.index:
                        text_value = limit_data.TEST_TXT
                    else:
                        continue

                    if text_value not in temp_lot.index:
                        continue
                    lot_limit_data = temp_lot.loc[text_value][each]  # 2022-1-20: CP25片用一个lot
                    sheet_1.cell(head_row + i, now_column).value = lot_limit_data
                    sheet_1.cell(head_row + i, now_column).alignment = OpenXl.CenterAlign
                    sheet_1.cell(head_row + i, now_column).font = OpenXl.TextFont
                    sheet_1.cell(head_row + i, now_column).border = OpenXl.TextBorder
                    if index != 0:
                        if sheet_1.cell(head_row + i, now_column).value != sheet_1.cell(head_row + i,
                                                                                        now_column - 2).value:
                            sheet_1.cell(head_row + i, now_column).fill = OpenXl.RedFill

                now_column += 1

        sheet_1.freeze_panes = "D{}".format(head_row)
        sheet_name = wb.worksheets[0]
        wb.remove(sheet_name)

        # 确保LIMIT_PATH目录存在
        if not os.path.exists(GlobalVariable.LIMIT_PATH):
            os.makedirs(GlobalVariable.LIMIT_PATH, exist_ok=True)

        try:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, 'limit.xlsx')
            wb.save(save_path)
        except:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, 'limit_{}.xlsx'.format(tid_maker()))
            wb.save(save_path)
        win32api.ShellExecute(0, 'open', save_path, '', '', 1)

    @staticmethod
    def excel_lot_compare_run(lot_df: pd.DataFrame, cpk_df: pd.DataFrame):
        """
        简单的数据, 使用快速透视的方法来写入到Excel中
        先尝试删除本地的数据, 如果删除不了, 就新建一个数据
        :return:
        """
        # 确保LIMIT_PATH目录存在
        if not os.path.exists(GlobalVariable.LIMIT_PATH):
            os.makedirs(GlobalVariable.LIMIT_PATH, exist_ok=True)

        save_path = os.path.join(GlobalVariable.LIMIT_PATH, "compare.xlsx")
        try:
            if os.path.exists(save_path):
                os.remove(save_path)
        except:
            save_path = os.path.join(GlobalVariable.LIMIT_PATH, "compare_{}.xlsx".format(tid_maker()))
        df = pd.merge(lot_df, cpk_df, on="ID")
        with pd.ExcelWriter(save_path) as writer:
            avg = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='AVG',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            std = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='STD',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            cpk = df.pivot_table(index='TEXT',  # 透视的行，分组依据
                                 columns=['SUB_CON', 'PART_TYP', 'JOB_NAM', 'TEST_COD', 'LOT_ID', 'SBLOT_ID',
                                          'WAFER_ID', 'QTY'],
                                 values='CPK',  # 值
                                 aggfunc='sum'  # 聚合函数
                                 )
            avg.to_excel(writer, 'AVG')
            std.to_excel(writer, 'STD')
            cpk.to_excel(writer, 'CPK')
        win32api.ShellExecute(0, 'open', save_path, '', '', 1)
